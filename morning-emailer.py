#!/usr/bin/env python

'''
Jeff's Morning Emailer
Jeffrey Neil Willits  @jnwillits
'''

import json
import os
import sys
from datetime import date
from os import listdir
from os.path import isfile, join
from pathlib import Path

import pyperclip
import PySimpleGUI as sg
import win32com.client as win32
from dateutil.parser import *
from openpyxl import load_workbook
from PIL import ImageGrab

about_info = """This is a utility to automate tasks associated with sending \
report emails that require attachments and information in the body of the \
email that is pasted from a specified Excel range. Put any email attachments \
in a subfolder named "attachments". The files will be automatically deleted \
when the email is sent. After you have completed the settings, you only need \
to press the "send" button. Entering your name to identify yourself as the \
sender is optional. This is the only setting that will not be stored by the \
program. This is not copyrighted and it is free for use. Python source code \
for this is available in my public GitHub repository.  Version 1.0 finished \
July 14, 2019.    Jeffrey Neil Willits, W: jnwillits.com \
"""

sg.ChangeLookAndFeel('Dark')
sg.SetOptions(icon='emailer_bpu_icon.ico', element_padding=(2, 7), font=('verdana', 8), text_color='#FFFAFA', button_color=('#FFFFFF', '#565656'),
              background_color='#1E1E1E', text_element_background_color='#1E1E1E')

email_dist = []
t1 = (150, 1)
t2 = (20, 1)
sent_by = ''

layout = [
    [sg.Button('', visible=False, ), ],
    [sg.Button('Excel File >', size=t2), sg.T(
        '', size=t1, key='_EXCEL_PATH_')],
    [sg.Button('Sheet Name >', size=t2), sg.T(
        '', size=t1, key='_SHEET_NAME_')],
    [sg.Button('Date Cell >', size=t2), sg.T('', size=t1, key='_DATE_CELL_')],
    [sg.Button('Day Cell >', size=t2), sg.T('', size=t1, key='_DAY_CELL_')],
    [sg.Button('Upper Cell >', size=t2), sg.T(
        '', size=t1, key='_UPPER_LEFT_CELL_')],
    [sg.Button('Lower Cell >', size=t2), sg.T(
        '', size=t1, key='_LOWER_RIGHT_CELL_')],
    [sg.Button('Attachments Folder >', size=t2), sg.T(
        '', size=t1, key='_ATTACHMENTS_FOLDER_')],
    [sg.Button('Subject >', size=t2), sg.T('', size=t1, key='_SUBJECT_')],
    [sg.T('')],
    [sg.T('Sent by (optional):', pad=(0, 0))],
    [sg.In(size=(30, 1), key='sent_by'), sg.T(
        '', size=(8, 1), key='sent_by_display'), ],
    [sg.T('')],
    [sg.T('Email Distribution:', pad=(0, 0)),   sg.T(
        '', size=(33, 0)),        sg.T('About:', pad=(0, 0)), ],
    [sg.Listbox(email_dist, change_submits=True, bind_return_key=True, size=(
        30, 7), key='_EMAILS_'), sg.T(''), sg.Multiline(about_info, size=(250, 1))],
    [sg.Button('Add Email', size=t2), sg.T('', size=t1, key='_ADD_EMAIL_')],
    [sg.Button('Remove Email', size=t2), sg.T('', size=(70, 1), key='_REMOVE_EMAIL_'), sg.Button('Send', size=t2), sg.T('', key='_SEND_')], ]


def delete_email_form(email_dist_pass):
    layout = [[sg.T('')],
              [sg.T('Select email to delete...', text_color='#FFFAFA')],
              [sg.Listbox(email_dist_pass, change_submits=True,
                          bind_return_key=True, size=(60, 5), key='_LIST2_')],
              [sg.T('')],
              [sg.Button('Delete', visible=True, key='_BUTTON_')], ]
    window_local = sg.Window('Delete an email...',
                             size=(300, 300)).Layout(layout)
    while True:
        event_local, values_local = window_local.Read(timeout=10)
        if event_local is None or event_local == 'Exit':
            break
        else:
            if event_local is not None:
                if event_local == '_BUTTON_':
                    email_selected = values_local['_LIST2_'][0]
                    if email_selected in email_dist_pass:
                        email_dist_pass.remove(email_selected)
                    window_local.Element('_LIST2_').Update(email_dist_pass)
    window_local.Close()
    return email_dist_pass


def save_data(setup_data_pass):
    with open('morning-emailer.json', 'w') as f_obj:
        json.dump(setup_data_pass, f_obj)


def get_spreadsheet_data(setup_data_pass):
    spreadsheet_data_local = []
    wb = load_workbook(
        filename=setup_data_pass['excel_file'], read_only=True, data_only=True)
    ws = wb[setup_data_pass['sheet']]
    date_obj = ws[setup_data_pass['date_cell']].value
    date_str = str(parse(str(date_obj)).date())
    date_str = date_str[5:7] + '/' + date_str[-2:] + '/' + date_str[:4]
    spreadsheet_data_local.append(ws[setup_data_pass['day_cell']].value)
    spreadsheet_data_local.append(date_str)
    return spreadsheet_data_local


def get_attachments(setup_data_pass):
    path_list = []
    file_names = [f for f in listdir(setup_data_pass['attachments_folder']) if isfile(
        join(setup_data_pass['attachments_folder'], f))]
    if len(file_names) > 0:
        for i in range(0, len(file_names)):
            path_list.append(
                setup_data_pass['attachments_folder'] + '/' + file_names[i])
    return path_list


def send_email(setup_data_pass, sent_by_pass, upper_cell_pass, lower_cell_pass, email_dist_pass):
    spreadsheet_data = get_spreadsheet_data(setup_data_pass)
    to_email = ''
    for i in range(0, len(email_dist_pass)):
        to_email += email_dist_pass[i] + ';'

    outlook = win32.gencache.EnsureDispatch('Outlook.Application')
    new_mail = outlook.CreateItem(0)
    subject_temp = setup_data['subject'].replace('[day]', spreadsheet_data[0])
    new_mail.Subject = subject_temp.replace('[date]', spreadsheet_data[1])

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(setup_data_pass['excel_file'])
    ws = wb.Worksheets(1)

    win32c = win32.constants
    range_str = f'{upper_cell_pass}:{lower_cell_pass}'
    ws.Range(range_str).CopyPicture(Format=win32c.xlBitmap)
    img = ImageGrab.grabclipboard()
    image_path = str(Path.cwd() / 'temp.png')
    img.save(image_path)

    attachments_list = get_attachments(setup_data)
    for i in range(0, len(attachments_list)):
        new_mail.Attachments.Add(Source=attachments_list[i])
        os.remove(attachments_list[i])

    new_mail.Attachments.Add(Source=image_path)
    html_str = f'<h5>&emsp;Sent by: {sent_by_pass}<h5><img src=temp.png>'
    body = html_str
    new_mail.HTMLBody = (body)

    wb.Close()
    new_mail.To = to_email
    new_mail.Send()
    sys.exit()


if __name__ == '__main__':
    var = ['excel_file', 'sheet', 'upper_cell', 'lower_cell', 'emails',
           'attachments_folder', 'date_cell', 'day_cell', 'subject']

    if os.path.isfile('morning-emailer.json'):
        with open('morning-emailer.json') as f_obj:
            setup_data = dict(json.load(f_obj))

    for i in range(len(var)):
        if var[i] not in setup_data.keys():
            setup_data.update({var[i]: ''})

    email_dist = setup_data['emails']

    window = sg.Window(" Jeff's Morning Emailer", size=(800, 640), default_element_size=(20, 1), grab_anywhere=False,
                       background_color='#1E1E1E', auto_size_text=False, auto_size_buttons=False).Layout(layout).Finalize()
    while True:
        event, values = window.Read(timeout=10)

        window.Element('_EXCEL_PATH_').Update(setup_data[var[0]])
        window.Element('_SHEET_NAME_').Update(setup_data[var[1]])
        window.Element('_DATE_CELL_').Update(setup_data[var[6]])
        window.Element('_DAY_CELL_').Update(setup_data[var[7]])
        window.Element('_UPPER_LEFT_CELL_').Update(setup_data[var[2]])
        window.Element('_LOWER_RIGHT_CELL_').Update(setup_data[var[3]])
        window.Element('_ADD_EMAIL_').Update()
        window.Element('_REMOVE_EMAIL_').Update()
        window.Element('_ATTACHMENTS_FOLDER_').Update(setup_data[var[5]])
        window.Element('_SUBJECT_').Update(setup_data[var[8]])
        window.Element('_EMAILS_').Update(setup_data[var[4]])
        window.Element('_SEND_').Update()
        window.Element('sent_by_display').Update(sent_by)

        f_path = setup_data[var[0]]

        if event is None or event == 'Exit':
            break
        else:
            if event == 'Excel File >':
                excel_file = sg.PopupGetFile('', 'Identify an Excel file...')
                setup_data.update({'excel_file': excel_file})
                save_data(setup_data)
            elif event == 'Sheet Name >':
                sheet_name = sg.PopupGetText(
                    '', 'Enter the Excel file sheet name.')
                setup_data.update({'sheet': sheet_name})
                save_data(setup_data)
            elif event == 'Date Cell >':
                date_cell = (sg.PopupGetText(
                    '', 'Enter Excel file date location.')).upper()
                setup_data.update({'date_cell': date_cell})
                save_data(setup_data)
            elif event == 'Day Cell >':
                day_cell = (sg.PopupGetText(
                    '', 'Enter Excel file day of week location.')).upper()
                setup_data.update({'day_cell': day_cell})
                save_data(setup_data)
            elif event == 'Upper Cell >':
                upper_cell = (sg.PopupGetText(
                    '', "Enter the upper left cell ('B3').")).upper()
                setup_data.update({'upper_cell': upper_cell})
                save_data(setup_data)
            elif event == 'Lower Cell >':
                lower_cell = (sg.PopupGetText(
                    '', "Enter the lower right cell ('E23').")).upper()
                setup_data.update({'lower_cell': lower_cell})
                save_data(setup_data)
            elif event == 'Add Email':
                new_email = (sg.PopupGetText(
                    '', 'Enter an email address to add.')).lower()
                email_dist.append(new_email)
                setup_data.update({'emails': email_dist})
                save_data(setup_data)
            elif event == 'Remove Email':
                email_dist = delete_email_form(email_dist)
                setup_data.update({'emails': email_dist})
                save_data(setup_data)
            if event == 'Attachments Folder >':
                attachments_folder = sg.PopupGetFolder(
                    '', 'Identify the folder for outgoing attachments.')
                setup_data.update({'attachments_folder': attachments_folder})
                save_data(setup_data)
            elif event == 'Subject >':
                subject = sg.PopupGetText('', "Enter the email subject line.")
                setup_data.update({'subject': subject})
                save_data(setup_data)
            elif event == 'Send':
                send_email(
                    setup_data, sent_by, setup_data['upper_cell'], setup_data['lower_cell'], email_dist)
        sent_by = values['sent_by']
        window.Element('sent_by_display').Update(sent_by)
    save_data(setup_data)
    window.close()
    sys.exit()
